"""
export_signal_board.py
Reads luck_scores.csv and pitcher_luck_scores.csv and writes
outputs/signal_board_YYYY-MM-DD.xlsx  (dated archive)
outputs/signal_board_latest.xlsx      (always-current share copy)

Usage:
    python export_signal_board.py
    python export_signal_board.py --out path/to/file.xlsx
"""

import math
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR   = Path(__file__).parent
HITTER_CSV = BASE_DIR / "luck_scores.csv"
PITCHER_CSV = BASE_DIR / "pitcher_luck_scores.csv"
H_RANK_CSV  = BASE_DIR / "data" / "fantasy_rankings_hitters_2026.csv"
P_RANK_CSV  = BASE_DIR / "data" / "fantasy_rankings_pitchers_2026.csv"
OUT_DIR     = BASE_DIR / "outputs"

# ---------------------------------------------------------------------------
# Color palette (ARGB — openpyxl stores as 8-char, first 2 = alpha)
# ---------------------------------------------------------------------------
C_DARK_NAVY   = "000D1117"   # title row background
C_HEADER_NAVY = "001E3A5F"   # column header background
C_LIGHT_GREY  = "00F9FAFB"   # subtitle background
C_ALT_ROW     = "00F9FAFB"   # alternating row
C_WHITE       = "00FFFFFF"
C_DARK_TEXT   = "00111827"
C_GREY_TEXT   = "006B7280"

# Per-tier signal cell colors
TIER_COLORS = {
    "Buy low":    ("00DCFCE7", "00166534"),   # green fill, green text
    "Slight buy": ("00DBEAFE", "001E40AF"),   # blue fill, blue text
    "Slight sell":("00FEF9C3", "00854D0E"),   # yellow fill, amber text
    "Sell high":  ("00FEE2E2", "00991B1B"),   # red fill, red text
}

C_POS_TEXT = "00166534"   # positive gap/luck text
C_NEG_TEXT = "00991B1B"   # negative gap/luck text

# This Is Real sheet colors (amber/gold theme)
C_REAL_TITLE  = "0092400E"
C_REAL_FILL   = "00FEF3C7"
C_REAL_CONF   = "0092400E"
C_REAL_MON    = "00B45309"

# This Is Actually Bad sheet colors (red/warning theme)
C_BAD_TITLE   = "007F1D1D"   # dark red title
C_BAD_CONF_F  = "00FEE2E2"   # red fill — confirmed
C_BAD_CONF_T  = "00991B1B"   # red text — confirmed
C_BAD_MON_F   = "00FEF9C3"   # amber fill — monitor
C_BAD_MON_T   = "00854D0E"   # amber text — monitor

BUY_VERDICTS  = {"Buy low", "Slight buy"}
SELL_VERDICTS = {"Sell high", "Slight sell"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)


def _font(rgb: str, bold: bool = False, size: float = 10) -> Font:
    return Font(color=rgb, bold=bold, size=size, name="Calibri")


def _align(h: str = "left") -> Alignment:
    return Alignment(horizontal=h, vertical="center")


def _norm(s: str) -> str:
    """Normalize name for fuzzy matching."""
    try:
        s = str(s).encode("latin1").decode("utf-8")
    except Exception:
        pass
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z ]", "", s.lower()).strip()


def _rank_lookup(csv_path: Path) -> dict:
    if not csv_path.exists():
        return {}
    try:
        df = pd.read_csv(csv_path)
        return {_norm(r["Player Name"]): int(r["Rank"]) for _, r in df.iterrows()}
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Formatters
# ---------------------------------------------------------------------------
def _stat(val, dec: int = 3) -> str:
    """Stat without leading zero: .241"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val)
    s = f"{abs(v):.{dec}f}"
    if s.startswith("0."):
        s = s[1:]
    return ("-" if v < 0 else "") + s


def _signed(val, dec: int = 3) -> str:
    """Gap with sign, no leading zero: +.044  or  -.031"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val)
    mag = _stat(abs(v), dec).lstrip("-")
    return f"+{mag}" if v >= 0 else f"-{mag}"


def _luck(val) -> str:
    """Luck score with leading zero: +0.426"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val)
    return f"+{v:.3f}" if v >= 0 else f"{v:.3f}"


def _era(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    return f"{float(val):.2f}"


def _era_gap(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    v = float(val)
    return f"+{v:.2f}" if v >= 0 else f"{v:.2f}"


def _pct(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    v = float(val)
    v = v * 100 if v <= 1.0 else v
    return f"{v:.1f}%"


def _own(val) -> str:
    """Ownership percentage: 67.3%"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return "-"
    return f"{v:.1f}%"


def _own_context(val) -> str:
    """Article context: 67.3% rostered"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return "-"
    return f"{v:.1f}% rostered"


def _bool_flag(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    return "Yes" if str(val).strip().lower() in ("true", "1", "yes") else "No"


def _age(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    try:
        return str(int(float(val)))
    except (TypeError, ValueError):
        return "-"


def _ip(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    return f"{float(val):.1f}"


def _text(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "-"
    return str(val)


# ---------------------------------------------------------------------------
# Signed color helper
# ---------------------------------------------------------------------------
def _sign_color(text: str) -> str | None:
    """Return a color code for signed values, or None for neutral."""
    if text and text[0] == "+":
        return C_POS_TEXT
    if text and text[0] == "-":
        return C_NEG_TEXT
    return None


# ---------------------------------------------------------------------------
# Sort signal rows: buy low desc -> slight buy desc -> slight sell asc -> sell high asc
# ---------------------------------------------------------------------------
TIER_ORDER = {"Buy low": 0, "Slight buy": 1, "Slight sell": 2, "Sell high": 3}
TIER_ASC   = {"Buy low": False, "Slight buy": False, "Slight sell": True, "Sell high": True}


def _sort_signals(df: pd.DataFrame, score_col: str = "luck_score") -> pd.DataFrame:
    active = df[df["verdict"].isin(BUY_VERDICTS | SELL_VERDICTS)].copy()
    has_own = "owned_pct" in active.columns and pd.to_numeric(active["owned_pct"], errors="coerce").notna().any()
    groups = []
    for tier in ["Buy low", "Slight buy", "Slight sell", "Sell high"]:
        grp = active[active["verdict"] == tier].copy()
        if has_own:
            grp = grp.sort_values("owned_pct", ascending=False, na_position="last")
        else:
            grp = grp.sort_values(score_col, ascending=TIER_ASC[tier])
        groups.append(grp)
    return pd.concat(groups, ignore_index=True)


# ---------------------------------------------------------------------------
# Sheet infrastructure
# ---------------------------------------------------------------------------
def _write_banner(ws, n_cols: int, title: str, subtitle: str,
                  buy_legend: str, sell_legend: str) -> None:
    last = get_column_letter(n_cols)

    # Row 1 — dark title
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(C_DARK_NAVY)
    c.font      = _font(C_WHITE, bold=True, size=13)
    c.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value     = subtitle
    c.fill      = _fill(C_LIGHT_GREY)
    c.font      = _font(C_GREY_TEXT, size=10)
    c.alignment = _align()
    ws.row_dimensions[2].height = 18

    # Row 3 — legend (3 merged sections)
    expl = ("Luck score: positive = underperforming (buy) | "
            "negative = overperforming (sell)")

    ws.merge_cells("A3:C3")
    c = ws["A3"]
    c.value     = buy_legend
    c.fill      = _fill(TIER_COLORS["Buy low"][0])
    c.font      = _font(TIER_COLORS["Buy low"][1], size=9)
    c.alignment = _align()

    ws.merge_cells("D3:F3")
    c = ws["D3"]
    c.value     = sell_legend
    c.fill      = _fill(TIER_COLORS["Sell high"][0])
    c.font      = _font(TIER_COLORS["Sell high"][1], size=9)
    c.alignment = _align()

    ws.merge_cells(f"G3:{last}3")
    c = ws["G3"]
    c.value     = expl
    c.font      = _font(C_GREY_TEXT, size=9)
    c.alignment = _align()

    ws.row_dimensions[3].height = 16


def _write_headers(ws, headers: list, widths: list) -> None:
    for i, (hdr, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(4, i)
        c.value     = hdr
        c.fill      = _fill(C_HEADER_NAVY)
        c.font      = _font(C_WHITE, bold=True, size=10)
        c.alignment = _align()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 20


def _write_row(ws, row_idx: int, values: list, verdict: str,
               signed_cols: set, signal_col: int) -> None:
    """
    Write one data row.
    signed_cols: 1-based col indices whose text should be green/red by sign.
    signal_col: 1-based index of the verdict/signal cell.
    """
    row_fill = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE

    for i, val in enumerate(values, 1):
        c = ws.cell(row_idx, i)
        c.value     = val
        c.alignment = _align()

        if i == signal_col:
            tf, tc = TIER_COLORS.get(verdict, (C_WHITE, C_DARK_TEXT))
            c.fill = _fill(tf)
            c.font = _font(tc, bold=True, size=10)
        elif i == 1:
            # Player name — bold, alternating background
            c.fill = _fill(row_fill)
            c.font = _font(C_DARK_TEXT, bold=True)
        elif i in signed_cols:
            sc = _sign_color(str(val) if val is not None else "")
            if sc:
                c.fill = _fill(row_fill)
                c.font = _font(sc, bold=True)
            else:
                c.fill = _fill(row_fill)
                c.font = _font(C_DARK_TEXT)
        else:
            c.fill = _fill(row_fill)
            c.font = _font(C_DARK_TEXT)

    ws.row_dimensions[row_idx].height = 18


# ---------------------------------------------------------------------------
# Hitter sheet
# ---------------------------------------------------------------------------
H_HEADERS = [
    "Player", "Team", "Rank", "Own%", "Article Context", "Signal",
    "wOBA", "xwOBA", "wOBA Gap", "BABIP", "Career BABIP", "BABIP Gap",
    "Luck Score", "Age", "Pattern",
]
H_WIDTHS = [22, 8, 8, 8, 18, 14, 9, 9, 10, 9, 14, 12, 12, 6, 26]
H_SIGNAL_COL  = 6    # "Signal" column index (1-based)
H_SIGNED_COLS = {9, 12, 13}   # wOBA Gap, BABIP Gap, Luck Score


def build_hitter_sheet(ws, df: pd.DataFrame, rank_lk: dict) -> None:
    today = date.today().strftime("%B %d, %Y")
    _write_banner(
        ws, len(H_HEADERS),
        title      = f"THE SIGNAL FANTASY - Hitter Signal Board | Updated: {today}",
        subtitle   = ("94.1% sell high accuracy  *  94.3% buy low accuracy  *  "
                      "4-year backtest 2022-2025  *  thesignalfantasy.substack.com"),
        buy_legend  = "Buy low / Slight buy",
        sell_legend = "Sell high / Slight sell",
    )
    _write_headers(ws, H_HEADERS, H_WIDTHS)

    active = _sort_signals(df)
    for offset, (_, r) in enumerate(active.iterrows()):
        row_idx = 5 + offset
        verdict = r.get("verdict", "")

        babip_gap_val = None
        try:
            b = float(r.get("BABIP", float("nan")))
            cb = float(r.get("career_babip", float("nan")))
            if not math.isnan(b) and not math.isnan(cb):
                babip_gap_val = b - cb
        except (TypeError, ValueError):
            pass

        name_key = _norm(r.get("name", ""))
        rank = rank_lk.get(name_key, "-")

        values = [
            _text(r.get("name")),
            _text(r.get("team", r.get("Team", ""))),
            rank,
            _own(r.get("owned_pct")),
            _own_context(r.get("owned_pct")),
            verdict,
            _stat(r.get("wOBA")),
            _stat(r.get("xwOBA")),
            _signed(r.get("xwOBA_gap")),
            _stat(r.get("BABIP")),
            _stat(r.get("career_babip")),
            _signed(babip_gap_val),
            _luck(r.get("luck_score")),
            _age(r.get("age")),
            _text(r.get("seasonal_pattern")),
        ]
        _write_row(ws, row_idx, values, verdict, H_SIGNED_COLS, H_SIGNAL_COL)

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Pitcher sheet
# ---------------------------------------------------------------------------
P_HEADERS = [
    "Player", "Team", "IP", "Own%", "Article Context", "Signal",
    "ERA", "FIP", "xERA", "ERA-FIP Gap",
    "BABIP Allowed", "Career BABIP", "LOB%",
    "Luck Score", "Age", "High Variance?", "LOB Signal", "April Pattern", "Evolution?",
]
P_WIDTHS = [22, 8, 7, 8, 18, 14, 8, 8, 8, 13, 14, 14, 8, 12, 6, 15, 12, 22, 16]
P_SIGNAL_COL  = 6
P_SIGNED_COLS = {10, 14}   # ERA-FIP Gap, Luck Score


def _evolution_label(row: pd.Series) -> str:
    """Return display label for pitcher evolution status."""
    score = row.get("evolution_score")
    if score is None or (isinstance(score, float) and score != score):
        return ""
    try:
        score = int(score)
    except (TypeError, ValueError):
        return ""
    if score >= 3:
        return "New pitcher"
    if score >= 1:
        return "Improving"
    if score <= -1:
        return "Declining"
    return ""


def build_pitcher_sheet(ws, df: pd.DataFrame, rank_lk: dict) -> None:
    today = date.today().strftime("%B %d, %Y")
    _write_banner(
        ws, len(P_HEADERS),
        title      = f"THE SIGNAL FANTASY - Pitcher Signal Board | Updated: {today}",
        subtitle   = ("94.6% sell high accuracy  *  82.1% buy low accuracy  *  "
                      "4-year backtest 2022-2025  *  thesignalfantasy.substack.com"),
        buy_legend  = "Buy low = ERA inflated by luck",
        sell_legend = "Sell high = ERA too good to last",
    )
    _write_headers(ws, P_HEADERS, P_WIDTHS)

    active = _sort_signals(df)
    for offset, (_, r) in enumerate(active.iterrows()):
        row_idx = 5 + offset
        verdict = r.get("verdict", "")

        evo_label = _evolution_label(r)
        evo_colors = {
            "New pitcher": ("00DBEAFE", "001E40AF"),  # blue
            "Improving":   ("00DCFCE7", "00166534"),  # green
            "Declining":   ("00FEE2E2", "00991B1B"),  # red
        }
        values = [
            _text(r.get("name")),
            _text(r.get("Team", "")),
            _ip(r.get("IP")),
            _own(r.get("owned_pct")),
            _own_context(r.get("owned_pct")),
            verdict,
            _era(r.get("ERA")),
            _era(r.get("FIP")),
            _era(r.get("xERA")),
            _era_gap(r.get("ERA_minus_FIP")),
            _stat(r.get("BABIP_allowed")),
            _stat(r.get("career_babip_allowed")),
            _pct(r.get("lob_pct")),
            _luck(r.get("luck_score")),
            _age(r.get("age")),
            _bool_flag(r.get("volatility_flag")),
            _bool_flag(r.get("lob_confluence_flag")),
            _text(r.get("april_pattern_flag")),
            evo_label,
        ]
        _write_row(ws, row_idx, values, verdict, P_SIGNED_COLS, P_SIGNAL_COL)

        # Color-code the Evolution column (last column)
        if evo_label and evo_label in evo_colors:
            evo_col = len(values)
            c = ws.cell(row_idx, evo_col)
            ef, et = evo_colors[evo_label]
            c.fill = _fill(ef)
            c.font = _font(et, bold=True)

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# This Is Real sheet
# ---------------------------------------------------------------------------
R_HEADERS = [
    "Player", "Team", "Status", "Own%", "Article Context",
    "wOBA", "xwOBA", "BABIP", "Career BABIP",
    "Luck Score", "Signal", "Age",
]
R_WIDTHS = [22, 8, 12, 8, 18, 9, 9, 9, 14, 12, 14, 6]
R_SIGNAL_COL  = 11
R_SIGNED_COLS = {10}   # Luck Score


def build_real_sheet(ws, df: pd.DataFrame) -> None:
    today = date.today().strftime("%B %d, %Y")
    n_cols = len(R_HEADERS)
    last = get_column_letter(n_cols)

    # Row 1 — amber title
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = f"THE SIGNAL FANTASY - Confirmed Performers | Updated: {today}"
    c.fill      = _fill("00451A03")
    c.font      = _font(C_WHITE, bold=True, size=13)
    c.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value = ("Players whose strong stats are backed by underlying contact quality - not luck  *  "
               "Confirmed = all quality gates pass  *  Monitor = one flag raised")
    c.fill      = _fill("00FFFBEB")
    c.font      = _font("00B45309", size=10)
    c.alignment = _align()
    ws.row_dimensions[2].height = 18

    # Row 3 — legend
    ws.merge_cells(f"A3:{last}3")
    c = ws["A3"]
    c.value     = ("Confirmed: wOBA >= .370, xwOBA >= .350, BABIP not inflated vs career, luck score > -0.050  |  "
                   "Monitor: strong wOBA but one quality flag raised")
    c.fill      = _fill("00FEF9C3")
    c.font      = _font("0092400E", size=9)
    c.alignment = _align()
    ws.row_dimensions[3].height = 16

    # Row 4 — headers
    for i, (hdr, w) in enumerate(zip(R_HEADERS, R_WIDTHS), 1):
        cell = ws.cell(4, i)
        cell.value     = hdr
        cell.fill      = _fill("00451A03")
        cell.font      = _font(C_WHITE, bold=True, size=10)
        cell.alignment = _align()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 20

    # Filter and sort: ownership desc (most rostered first for article triage)
    real = df[df["this_is_real"].isin(["Confirmed", "Monitor"])].copy()
    if "owned_pct" in real.columns:
        real = real.sort_values("owned_pct", ascending=False, na_position="last")
    else:
        real = real.sort_values("wOBA", ascending=False)

    for offset, (_, r) in enumerate(real.iterrows()):
        row_idx = 5 + offset
        status  = r.get("this_is_real", "")
        verdict = r.get("verdict", "Neutral")
        row_fill = C_ALT_ROW if row_idx % 2 == 0 else "00FFFBEB"

        # Status cell color
        if status == "Confirmed":
            status_fill, status_fc = "00FEF3C7", "0092400E"
        else:
            status_fill, status_fc = "00FFFBEB", "00B45309"

        values = [
            _text(r.get("name")),
            _text(r.get("team", r.get("Team", ""))),
            status,
            _own(r.get("owned_pct")),
            _own_context(r.get("owned_pct")),
            _stat(r.get("wOBA")),
            _stat(r.get("xwOBA")),
            _stat(r.get("BABIP")),
            _stat(r.get("career_babip")),
            _luck(r.get("luck_score")),
            verdict,
            _age(r.get("age")),
        ]

        for i, val in enumerate(values, 1):
            c = ws.cell(row_idx, i)
            c.value     = val
            c.alignment = _align()

            if i == 1:
                c.fill = _fill(row_fill)
                c.font = _font(C_DARK_TEXT, bold=True)
            elif i == 3:   # Status
                c.fill = _fill(status_fill)
                c.font = _font(status_fc, bold=True)
            elif i == R_SIGNAL_COL:   # Signal/verdict
                tf, tc = TIER_COLORS.get(verdict, (C_WHITE, C_DARK_TEXT))
                c.fill = _fill(tf)
                c.font = _font(tc, bold=True)
            elif i in R_SIGNED_COLS:
                sc = _sign_color(str(val) if val is not None else "")
                if sc:
                    c.fill = _fill(row_fill)
                    c.font = _font(sc, bold=True)
                else:
                    c.fill = _fill(row_fill)
                    c.font = _font(C_DARK_TEXT)
            else:
                c.fill = _fill(row_fill)
                c.font = _font(C_DARK_TEXT)

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# This Is Actually Bad sheet
# ---------------------------------------------------------------------------
B_HEADERS = [
    "Player", "Team", "Status", "Ownership Tier", "Own%", "Article Context",
    "wOBA", "xwOBA", "BABIP", "Career BABIP",
    "Luck Score", "Signal", "Age",
]
B_WIDTHS      = [22, 8, 12, 18, 8, 18, 9, 9, 9, 14, 12, 14, 6]
B_SIGNAL_COL  = 12    # "Signal" column (1-based)
B_SIGNED_COLS = {11}  # Luck Score


def build_bad_sheet(ws, df: pd.DataFrame) -> None:
    today = date.today().strftime("%B %d, %Y")
    n_cols = len(B_HEADERS)
    last = get_column_letter(n_cols)

    # Row 1 — dark red title
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = f"THE SIGNAL FANTASY — Confirmed Underperformers | Updated: {today}"
    c.fill      = _fill(C_BAD_TITLE)
    c.font      = _font(C_WHITE, bold=True, size=13)
    c.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value = ("Players whose poor stats are backed by weak contact quality — not bad luck. "
               "Consider your exit options.")
    c.fill      = _fill(C_BAD_CONF_F)
    c.font      = _font(C_BAD_CONF_T, size=10)
    c.alignment = _align()
    ws.row_dimensions[2].height = 18

    # Row 3 — legend
    ws.merge_cells(f"A3:{last}3")
    c = ws["A3"]
    c.value     = ("Confirmed = poor stats backed by weak contact. "
                   "Monitor = struggling but less certain. "
                   "Use ownership tier to assess trade value.")
    c.fill      = _fill(C_BAD_MON_F)
    c.font      = _font(C_BAD_MON_T, size=9)
    c.alignment = _align()
    ws.row_dimensions[3].height = 16

    # Row 4 — headers
    for i, (hdr, w) in enumerate(zip(B_HEADERS, B_WIDTHS), 1):
        cell = ws.cell(4, i)
        cell.value     = hdr
        cell.fill      = _fill(C_BAD_TITLE)
        cell.font      = _font(C_WHITE, bold=True, size=10)
        cell.alignment = _align()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 20

    if "this_is_actually_bad" not in df.columns:
        ws["A5"] = "Run score_luck.py to generate this_is_actually_bad column."
        return

    has_own = "ownership_tier" in df.columns

    confirmed = df[df["this_is_actually_bad"] == "Confirmed"].copy()
    monitor   = df[df["this_is_actually_bad"] == "Monitor"].copy()

    # Monitor filtered to rostered players only — fringe Monitor is not actionable
    if has_own:
        monitor = monitor[monitor["ownership_tier"].isin(["Widely rostered", "Commonly rostered"])]

    if "owned_pct" in confirmed.columns:
        confirmed = confirmed.sort_values("owned_pct", ascending=False, na_position="last")
        monitor   = monitor.sort_values("owned_pct", ascending=False, na_position="last")
    else:
        confirmed = confirmed.sort_values("wOBA", ascending=True)
        monitor   = monitor.sort_values("wOBA", ascending=True)

    display = pd.concat([confirmed, monitor], ignore_index=True)

    for offset, (_, r) in enumerate(display.iterrows()):
        row_idx  = 5 + offset
        status   = r.get("this_is_actually_bad", "")
        verdict  = r.get("verdict", "Neutral")
        row_fill = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE

        if status == "Confirmed":
            s_fill, s_fc = C_BAD_CONF_F, C_BAD_CONF_T
        else:
            s_fill, s_fc = C_BAD_MON_F, C_BAD_MON_T

        values = [
            _text(r.get("name")),
            _text(r.get("team", r.get("Team", ""))),
            status,
            _text(r.get("ownership_tier", "Fringe") if has_own else "-"),
            _own(r.get("owned_pct")),
            _own_context(r.get("owned_pct")),
            _stat(r.get("wOBA")),
            _stat(r.get("xwOBA")),
            _stat(r.get("BABIP")),
            _stat(r.get("career_babip")),
            _luck(r.get("luck_score")),
            verdict,
            _age(r.get("age")),
        ]

        for i, val in enumerate(values, 1):
            c = ws.cell(row_idx, i)
            c.value     = val
            c.alignment = _align()

            if i == 1:
                c.fill = _fill(row_fill)
                c.font = _font(C_DARK_TEXT, bold=True)
            elif i == 3:   # Status
                c.fill = _fill(s_fill)
                c.font = _font(s_fc, bold=True)
            elif i == B_SIGNAL_COL:   # Signal
                tf, tc = TIER_COLORS.get(verdict, (C_WHITE, C_DARK_TEXT))
                c.fill = _fill(tf)
                c.font = _font(tc, bold=True)
            elif i in B_SIGNED_COLS:
                sc = _sign_color(str(val) if val is not None else "")
                if sc:
                    c.fill = _fill(row_fill)
                    c.font = _font(sc, bold=True)
                else:
                    c.fill = _fill(row_fill)
                    c.font = _font(C_DARK_TEXT)
            else:
                c.fill = _fill(row_fill)
                c.font = _font(C_DARK_TEXT)

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# How to Use sheet
# ---------------------------------------------------------------------------
HOW_TO_USE = [
    (1,  "HOW TO USE THE SIGNAL FANTASY SIGNAL BOARD", None, True, 12),
    (3,  "WHAT IS THIS?", None, True, 10),
    (4,  None,
     "The Signal Fantasy uses 7 layers of Statcast data to identify hitters and pitchers "
     "whose stats are misleading - either better or worse than their underlying performance warrants.",
     False, 9),
    (5,  None,
     "Buy low = player is underperforming their true talent. Stats look bad, underlying data "
     "says regression upward is coming.",
     False, 9),
    (7,  "KEY METRICS", None, True, 10),
    (8,  "wOBA",          "Weighted On-Base Average - a comprehensive measure of offensive value.", False, 9),
    (9,  "xwOBA",         "Expected wOBA based on exit velocity and launch angle - what the player deserved based on contact quality.", False, 9),
    (10, "wOBA Gap",      "xwOBA minus actual wOBA. Positive = underperforming contact quality (buy indicator).", False, 9),
    (11, "BABIP",         "Batting Average on Balls in Play - highly luck-dependent, mean-reverts toward career baseline.", False, 9),
    (12, "Career BABIP",  "This player's own historical BABIP baseline (NOT league average). Individualizes the luck threshold.", False, 9),
    (13, "BABIP Gap",     "Current BABIP minus career baseline. Negative = running below own baseline (buy signal).", False, 9),
    (14, "ERA-FIP Gap",   "ERA minus FIP. Positive = ERA inflated by luck. FIP strips out defense and sequencing.", False, 9),
    (15, "xERA",          "Expected ERA based on Statcast contact quality. Confirms or challenges the FIP signal.", False, 9),
    (16, "LOB%",          "Strand rate. League average ~72%. Above 80% = unsustainable luck for pitchers.", False, 9),
    (17, "Luck Score",    "Model's composite signal. Positive = buy. Negative = sell. Magnitude = confidence.", False, 9),
    (18, "High Variance?","Pitcher with inconsistent start quality. Signals less reliable for high-variance pitchers.", False, 9),
    (19, "April Pattern", "Multi-year April trend - identifies pitchers who historically show April luck distortions.", False, 9),
    (21, "HOW TO USE IT", None, True, 10),
    (22, "Buy low",       "Player's stats look bad but underlying data says talent is there. Buy before the rebound.", False, 9),
    (23, "Slight buy",    "Moderate signal. Stats slightly below true talent. Lower confidence, smaller move.", False, 9),
    (24, "Slight sell",   "Moderate signal. Stats slightly above true talent. Trade at value if you can.", False, 9),
    (25, "Sell high",     "Strong signal. Stats well above true talent. Sell before regression hits.", False, 9),
    (27, "Confirmed Performer",
     "Strong actual stats (wOBA >= .370) backed by contact quality (xwOBA >= .350) with normal BABIP. Performance is real.",
     False, 9),
    (28, "Monitor",       "Strong wOBA but one flag raised (elevated BABIP or weaker contact backing). Watch for regression.", False, 9),
    (29, "This Is Actually Bad (Confirmed)",
     "Poor wOBA (.280 or lower) backed by weak contact quality. BABIP is not unusually low — the bad stats are real. "
     "Consider selling at face value before the market catches up.",
     False, 9),
    (30, "This Is Actually Bad (Monitor)",
     "Struggling player where contact quality mostly confirms poor performance. One flag of uncertainty remains. "
     "Monitor before acting.",
     False, 9),
    (31, "Ownership Tier",
     "Widely rostered (rank 1-150), Commonly rostered (151-300), Deep league (301-500), Fringe (500+). "
     "Use this to assess whether a trade market exists for the player.",
     False, 9),
    (33, "ACCURACY",      None, True, 10),
    (34, "Hitters",       "94.1% sell high  *  94.3% buy low  *  88.1% overall  *  +17.9pp vs regression-to-mean baseline", False, 9),
    (35, "Pitchers",      "94.6% sell high  *  82.1% buy low  *  91.1% overall  *  +21.1pp vs regression-to-mean baseline", False, 9),
    (36, "Validation",    "Tested on 2025 data the model never trained on. Track record since 2022.", False, 9),
    (38, "IMPORTANT CAVEATS", None, True, 10),
    (39, None, "Signals are directional, not prescriptive. Whether to trade depends on your league context.", False, 9),
    (40, None, "Buy low on a $1 keeper is different from buy low in a redraft - use your judgment.", False, 9),
    (41, None, "Injuries, platoon situations, and park changes can affect signal reliability.", False, 9),
    (43, "CONTACT",       "thesignalfantasy@gmail.com  |  @SignalFantasy  |  thesignalfantasy.substack.com", False, 9),
]


def build_how_to_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75

    for row_idx, label, body, bold, size in HOW_TO_USE:
        if label and body:
            c1 = ws.cell(row_idx, 1)
            c1.value     = label
            c1.font      = _font(C_DARK_TEXT, bold=True, size=size)
            c1.alignment = _align()
            c2 = ws.cell(row_idx, 2)
            c2.value     = body
            c2.font      = _font(C_DARK_TEXT, size=size)
            c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif label:
            c = ws.cell(row_idx, 1)
            c.value     = label
            c.font      = _font(C_DARK_TEXT, bold=bold, size=size)
            c.alignment = _align()
        elif body:
            c = ws.cell(row_idx, 2)
            c.value     = body
            c.font      = _font(C_DARK_TEXT, size=size)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 18 if size <= 9 else 22


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------
def export(output_path: Path | None = None) -> Path:
    if not HITTER_CSV.exists():
        raise FileNotFoundError(f"Missing: {HITTER_CSV}  — run score_luck.py first")
    if not PITCHER_CSV.exists():
        raise FileNotFoundError(f"Missing: {PITCHER_CSV}  — run score_pitcher_luck.py first")

    hitters  = pd.read_csv(HITTER_CSV)
    pitchers = pd.read_csv(PITCHER_CSV)

    rank_h = _rank_lookup(H_RANK_CSV)
    rank_p = _rank_lookup(P_RANK_CSV)

    # Build workbook
    wb = openpyxl.Workbook()
    ws_h = wb.active
    ws_h.title = "Hitter Signals"
    ws_p = wb.create_sheet("Pitcher Signals")
    ws_r = wb.create_sheet("This Is Real")
    ws_b = wb.create_sheet("This Is Actually Bad")
    ws_u = wb.create_sheet("How to Use This Sheet")

    build_hitter_sheet(ws_h, hitters, rank_h)
    build_pitcher_sheet(ws_p, pitchers, rank_p)

    if "this_is_real" in hitters.columns:
        build_real_sheet(ws_r, hitters)
    else:
        ws_r["A1"] = "Run score_luck.py to generate this_is_real column."

    build_bad_sheet(ws_b, hitters)

    build_how_to_sheet(ws_u)

    # Save dated + latest
    OUT_DIR.mkdir(exist_ok=True)
    today_str = date.today().isoformat()

    if output_path is None:
        dated_path  = OUT_DIR / f"signal_board_{today_str}.xlsx"
        latest_path = OUT_DIR / "signal_board_latest.xlsx"
    else:
        dated_path  = Path(output_path)
        latest_path = dated_path.parent / "signal_board_latest.xlsx"

    wb.save(dated_path)
    shutil.copy(dated_path, latest_path)

    # Summary
    h_active = hitters[hitters["verdict"].isin(BUY_VERDICTS | SELL_VERDICTS)]
    p_active = pitchers[pitchers["verdict"].isin(BUY_VERDICTS | SELL_VERDICTS)]
    h_real   = hitters[hitters.get("this_is_real", pd.Series()).isin(["Confirmed", "Monitor"])] \
               if "this_is_real" in hitters.columns else pd.DataFrame()

    hvc = h_active["verdict"].value_counts().to_dict()
    pvc = p_active["verdict"].value_counts().to_dict()

    print(f"  Signal board saved -> {dated_path.name}")
    print(f"  Latest copy      -> {latest_path.name}")
    print(f"  Hitters : {hvc.get('Buy low',0)} buy low | "
          f"{hvc.get('Slight buy',0)} slight buy | "
          f"{hvc.get('Slight sell',0)} slight sell | "
          f"{hvc.get('Sell high',0)} sell high")
    print(f"  Pitchers: {pvc.get('Buy low',0)} buy low | "
          f"{pvc.get('Slight buy',0)} slight buy | "
          f"{pvc.get('Slight sell',0)} slight sell | "
          f"{pvc.get('Sell high',0)} sell high")
    if not h_real.empty:
        real_counts = h_real.get("this_is_real", pd.Series()).value_counts().to_dict() if "this_is_real" in h_real.columns else {}
        print(f"  This Is Real: {real_counts.get('Confirmed',0)} confirmed | "
              f"{real_counts.get('Monitor',0)} monitor")

    if "this_is_actually_bad" in hitters.columns:
        h_bad = hitters[hitters["this_is_actually_bad"].isin(["Confirmed", "Monitor"])]
        has_own = "ownership_tier" in hitters.columns
        monitor_shown = (
            h_bad[
                (h_bad["this_is_actually_bad"] == "Monitor") &
                h_bad["ownership_tier"].isin(["Widely rostered", "Commonly rostered"])
            ] if has_own else h_bad[h_bad["this_is_actually_bad"] == "Monitor"]
        )
        print(f"  This Is Actually Bad: {(h_bad['this_is_actually_bad']=='Confirmed').sum()} confirmed "
              f"(all shown) | {len(monitor_shown)} monitor shown (rostered only)")

    return latest_path


if __name__ == "__main__":
    out = None
    if "--out" in sys.argv:
        idx = sys.argv.index("--out")
        if idx + 1 < len(sys.argv):
            out = Path(sys.argv[idx + 1])
    export(out)
